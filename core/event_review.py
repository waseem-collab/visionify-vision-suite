#!/usr/bin/env python3
"""
Event Review — raw vs inference, side by side.

Port of the standalone Desktop/check app into the suite:

1. Upload a CSV / TSV / XLSX export of events.
2. Refresh every Azure SAS URL in the sheet in one pass (suite credentials).
3. Extract frame 24 of every raw + inference video in parallel — via the
   Frame Extractor's download-then-decode path (measured ~25x faster than
   streaming ffmpeg over HTTPS).
4. Review raw vs inference one event at a time; tag events; export the sheet
   with a review_tags column.
5. Play either video on demand (downloaded once, cached).

Sessions are keyed by the upload's content hash and live under
``paths.EVENT_REVIEW_ROOT/<sid>/`` (frames/, videos/, tags.json), so
re-uploading the same export — even after a restart — resumes with everything
already extracted. Flask routes in run.py delegate here.
"""
import csv
import hashlib
import io
import json
import logging
import re
import threading
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Dict, List, Optional
from urllib.parse import urlsplit

from core import frame_extract, paths

LOGGER = logging.getLogger("eventreview")

FRAME_NUMBER = 24
FRAME_WORKERS = 10
VIDEO_WORKERS = 4

CSV_EXT = {".csv", ".tsv", ".txt"}
EXCEL_EXT = {".xlsx", ".xlsm"}

RAW_HINTS = ["raw_video_disk_path", "raw_video", "raw_path", "raw_url", "raw"]
INF_HINTS = ["inference_video_path", "inference_video", "inference_path",
             "inference_url", "inference"]

BLOB_RE = re.compile(r"^https?://[^\s]*\.blob\.core\.windows\.net/", re.I)
MEDIA_EXT = (".mp4", ".avi", ".mkv", ".mov", ".webm", ".jpg", ".jpeg", ".png")

ROOT = paths.EVENT_REVIEW_ROOT
UPLOAD_DIR = ROOT / "uploads"


class ApiError(Exception):
    def __init__(self, code, message):
        super().__init__(message)
        self.code = code
        self.message = message


# --------------------------------------------------------------------------- #
# Table loading
# --------------------------------------------------------------------------- #
def _strip_leading_blanks(text: str) -> str:
    lines = text.splitlines(True)
    i = 0
    while i < len(lines) and not lines[i].strip():
        i += 1
    return "".join(lines[i:])


def load_table(path: Path):
    """Return (headers, rows) for a csv/tsv/txt/xlsx file."""
    suffix = path.suffix.lower()
    if suffix in EXCEL_EXT:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = None
        rows = []
        for r in ws.iter_rows(values_only=True):
            values = ["" if c is None else str(c).strip() for c in r]
            if headers is None:
                if not any(values):
                    continue
                headers = [v.lstrip("﻿").strip() for v in values]
                continue
            if any(values):
                rows.append(values)
        wb.close()
        if headers is None:
            raise ValueError("The sheet appears to be empty.")
        return headers, rows

    if suffix not in CSV_EXT:
        raise ValueError(f"Unsupported file type '{suffix}'. "
                         f"Use one of: {', '.join(sorted(CSV_EXT | EXCEL_EXT))}")

    text = _strip_leading_blanks(path.read_text(encoding="utf-8-sig", errors="replace"))
    if not text.strip():
        raise ValueError("The file appears to be empty.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
    except csv.Error:
        first = sample.splitlines()[0] if sample.splitlines() else ""
        if first.count("\t") > first.count(","):
            class _T(csv.excel):
                delimiter = "\t"
            dialect = _T()
        else:
            dialect = csv.excel()

    reader = csv.reader(io.StringIO(text), dialect=dialect)
    headers = [h.lstrip("﻿").strip() for h in next(reader)]
    rows = [r for r in reader if any(c.strip() for c in r)]
    width = len(headers)
    rows = [(r + [""] * width)[:width] for r in rows]
    return headers, rows


def guess_column(headers: List[str], hints: List[str], exclude: Optional[int] = None):
    low = [h.lower() for h in headers]
    for hint in hints:
        for i, h in enumerate(low):
            if i != exclude and h == hint:
                return i
    for hint in hints:
        for i, h in enumerate(low):
            if i != exclude and hint in h and ("video" in h or "path" in h or "url" in h):
                return i
    for hint in hints:
        for i, h in enumerate(low):
            if i != exclude and hint in h:
                return i
    return None


def looks_like_blob_url(value: str) -> bool:
    return bool(value) and bool(BLOB_RE.match(value.strip()))


def looks_like_blob_path(value: str) -> bool:
    """A bare container-relative blob path (some rows store these instead of a
    full URL), e.g. 'ANSA-McAL/.../videos/raw/Camera_20260820-064617.mp4'."""
    v = (value or "").strip()
    if not v or v.startswith("#") or "://" in v or " " in v or "/" not in v:
        return False
    return v.lower().endswith(MEDIA_EXT)


# --------------------------------------------------------------------------- #
# Session state
# --------------------------------------------------------------------------- #
class Session:
    def __init__(self, sid: str, filename: str, headers: List[str], rows: List[List[str]]):
        self.id = sid
        self.filename = filename
        self.headers = headers
        self.rows = rows
        self.dir = ROOT / sid
        self.frames_dir = self.dir / "frames"
        self.videos_dir = self.dir / "videos"
        for d in (self.frames_dir, self.videos_dir):
            d.mkdir(parents=True, exist_ok=True)

        self.raw_col: Optional[int] = None
        self.inf_col: Optional[int] = None

        self.lock = threading.Lock()
        self.refresh_state = {"state": "idle", "done": 0, "total": 0, "updated": 0, "failed": 0}
        self.frame_state = {"state": "idle", "done": 0, "total": 0, "ok": 0, "failed": 0}
        self.frame_status: Dict[str, str] = {}
        self.video_status: Dict[str, dict] = {}
        self.abort = threading.Event()
        self.url_cols: List[int] = []
        self.constant_cols: Dict[str, str] = {}
        self.varying_cols: List[int] = []
        self.container: Optional[str] = self.detect_container()
        self.classify_columns()
        self.video_pool = ThreadPoolExecutor(max_workers=VIDEO_WORKERS,
                                             thread_name_prefix="ev-video")
        self.tags_file = self.dir / "tags.json"
        self.tags: Dict[int, List[str]] = {}
        self.load_tags()

    # -- review tags ------------------------------------------------------- #
    def load_tags(self):
        try:
            if self.tags_file.exists():
                raw = json.loads(self.tags_file.read_text(encoding="utf-8"))
                self.tags = {int(k): list(v) for k, v in raw.items() if v}
        except Exception as exc:
            LOGGER.warning("Could not read %s: %s", self.tags_file, exc)

    def save_tags(self):
        tmp = self.tags_file.with_suffix(".tmp")
        tmp.write_text(json.dumps({str(k): v for k, v in sorted(self.tags.items())},
                                  indent=1), encoding="utf-8")
        tmp.replace(self.tags_file)

    def tag_counts(self) -> List[dict]:
        counts: Dict[str, int] = {}
        for names in self.tags.values():
            for n in names:
                counts[n] = counts.get(n, 0) + 1
        return [{"name": n, "count": c}
                for n, c in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0].lower()))]

    def detect_container(self) -> Optional[str]:
        """Container name from the first full blob URL, so bare container-
        relative paths in other rows can be rebuilt into real URLs."""
        for row in self.rows:
            for v in row:
                v = v.strip()
                if looks_like_blob_url(v):
                    path = urlsplit(v).path.lstrip("/")
                    if "/" in path:
                        return path.split("/", 1)[0]
        return None

    def classify_columns(self):
        """Split columns into URL columns, constants (shown once) and columns
        that actually vary; drop the empty ones. Keeps the per-event details
        readable instead of a 60-field dump of placeholders."""
        for c, name in enumerate(self.headers):
            seen = set()
            nonempty = 0
            urlish = 0
            for row in self.rows:
                v = row[c].strip() if c < len(row) else ""
                if not v:
                    continue
                nonempty += 1
                if v.lower().startswith("http") or looks_like_blob_path(v):
                    urlish += 1
                if len(seen) < 3:
                    seen.add(v)
            if nonempty == 0:
                continue
            if urlish and urlish >= nonempty * 0.8:
                self.url_cols.append(c)
            elif len(seen) == 1 and nonempty == len(self.rows):
                self.constant_cols[name] = next(iter(seen))
            else:
                self.varying_cols.append(c)

    # -- helpers ----------------------------------------------------------- #
    def url(self, idx: int, kind: str) -> str:
        col = self.raw_col if kind == "raw" else self.inf_col
        if col is None or idx < 0 or idx >= len(self.rows):
            return ""
        return self.rows[idx][col].strip()

    def frame_path(self, idx: int, kind: str) -> Path:
        return self.frames_dir / f"{idx:06d}_{kind}.jpg"

    def video_path(self, idx: int, kind: str) -> Path:
        return self.videos_dir / f"{idx:06d}_{kind}.mp4"


SESSIONS: Dict[str, Session] = {}
CURRENT: Dict[str, str] = {"sid": ""}


def get_session(sid: Optional[str] = None) -> Session:
    sid = sid or CURRENT["sid"]
    s = SESSIONS.get(sid)
    if not s:
        raise ApiError(404, "No active session. Upload a file first.")
    return s


def create_session(filename: str, data: bytes):
    """Store the upload (content-hash keyed) and build/reuse its session."""
    name = Path(filename or "upload.csv").name
    suffix = Path(name).suffix.lower()
    if suffix not in CSV_EXT | EXCEL_EXT:
        raise ApiError(400, f"Unsupported file type '{suffix}'. "
                            f"Upload {', '.join(sorted(CSV_EXT | EXCEL_EXT))}.")
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    sid = hashlib.sha256(data).hexdigest()[:12]
    dest = UPLOAD_DIR / f"{sid}_{name}"
    dest.write_bytes(data)

    existing = SESSIONS.get(sid)
    if existing is not None:
        CURRENT["sid"] = sid
        blob_urls = sum(1 for row in existing.rows for v in row if looks_like_blob_url(v))
        return {"session_id": sid, "filename": existing.filename,
                "rows": len(existing.rows), "columns": existing.headers,
                "raw_col": existing.raw_col, "inference_col": existing.inf_col,
                "blob_urls": blob_urls, "resumed": True}

    try:
        headers, rows = load_table(dest)
    except Exception as exc:
        raise ApiError(400, f"Could not read the file: {exc}")
    if not rows:
        raise ApiError(400, "The file has a header row but no data rows.")

    s = Session(sid, name, headers, rows)
    s.raw_col = guess_column(headers, RAW_HINTS)
    s.inf_col = guess_column(headers, INF_HINTS, exclude=s.raw_col)
    SESSIONS[sid] = s
    CURRENT["sid"] = sid

    cached = len(list(s.frames_dir.glob("*.jpg")))
    blob_urls = sum(1 for row in rows for v in row if looks_like_blob_url(v))
    return {"session_id": sid, "cached_frames": cached, "filename": name,
            "rows": len(rows), "columns": headers,
            "raw_col": s.raw_col, "inference_col": s.inf_col,
            "blob_urls": blob_urls}


# --------------------------------------------------------------------------- #
# SAS refresh — whole sheet, one pass (suite Azure credentials)
# --------------------------------------------------------------------------- #
def start_refresh(s: Session):
    if s.refresh_state.get("state") == "running":
        return
    try:
        frame_extract._sas_credentials()
    except Exception as exc:
        raise ApiError(500, str(exc))
    threading.Thread(target=_refresh_all_urls, args=(s,), daemon=True).start()


def _refresh_all_urls(s: Session):
    account_name, account_key = frame_extract._sas_credentials()
    host = f"https://{account_name}.blob.core.windows.net"
    # container from the sheet's own full URLs, else the account's default —
    # so a sheet of ONLY bare paths still promotes into working URLs
    container = s.container or frame_extract.default_container()
    targets = []
    promote = 0
    for r, row in enumerate(s.rows):
        for c, value in enumerate(row):
            if looks_like_blob_url(value):
                targets.append((r, c))
            elif looks_like_blob_path(value):
                s.rows[r][c] = f"{host}/{container}/{value.strip().lstrip('/')}"
                targets.append((r, c))
                promote += 1

    with s.lock:
        s.refresh_state = {"state": "running", "done": 0, "total": len(targets),
                           "updated": 0, "failed": 0, "rebuilt": promote}
    if not targets:
        with s.lock:
            s.refresh_state["state"] = "done"
        return

    def work(item):
        r, c = item
        try:
            new = frame_extract.refresh_sas_url(s.rows[r][c].strip(),
                                                account_name, account_key)
        except Exception:
            new = None
        with s.lock:
            s.refresh_state["done"] += 1
            if new:
                if new != s.rows[r][c]:
                    s.rows[r][c] = new
                    s.refresh_state["updated"] += 1
            else:
                s.refresh_state["failed"] += 1

    try:
        with ThreadPoolExecutor(max_workers=16, thread_name_prefix="ev-sas") as pool:
            list(pool.map(work, targets))
        out = s.dir / f"{Path(s.filename).stem}_refreshed.csv"
        with out.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(s.headers)
            w.writerows(s.rows)
        with s.lock:
            s.refresh_state["state"] = "done"
            s.refresh_state["output"] = str(out)
    except Exception as exc:
        LOGGER.exception("SAS refresh failed")
        with s.lock:
            s.refresh_state["state"] = "error"
            s.refresh_state["error"] = str(exc)


# --------------------------------------------------------------------------- #
# Frame extraction (parallel, download-then-decode)
# --------------------------------------------------------------------------- #
def _extract_frame(url: str, out: Path) -> bool:
    import cv2
    try:
        frame = frame_extract._grab_frame(url, FRAME_NUMBER)
    except Exception:
        return False
    if frame is None:
        return False
    tmp = out.with_suffix(".part.jpg")
    if not cv2.imwrite(str(tmp), frame):
        tmp.unlink(missing_ok=True)
        return False
    tmp.replace(out)
    return True


def start_frames(s: Session):
    if s.raw_col is None or s.inf_col is None:
        raise ApiError(400, "Raw and inference columns are not set.")
    if s.frame_state.get("state") == "running":
        return
    s.abort.clear()

    tasks = []
    for idx in range(len(s.rows)):
        for kind in ("raw", "inference"):
            if s.url(idx, kind).lower().startswith("http"):
                tasks.append((idx, kind))

    with s.lock:
        s.frame_state = {"state": "running", "done": 0, "total": len(tasks),
                         "ok": 0, "failed": 0}
        for idx, kind in tasks:
            s.frame_status[f"{idx}_{kind}"] = "pending"

    def work(task):
        idx, kind = task
        if s.abort.is_set():
            with s.lock:
                s.frame_state["done"] += 1
            return
        out = s.frame_path(idx, kind)
        ok = out.exists() and out.stat().st_size > 0
        if not ok:
            ok = _extract_frame(s.url(idx, kind), out)
        with s.lock:
            s.frame_state["done"] += 1
            s.frame_state["ok" if ok else "failed"] += 1
            s.frame_status[f"{idx}_{kind}"] = "ok" if ok else "failed"

    def runner():
        try:
            # submitted in row order so the earliest events become reviewable first
            with ThreadPoolExecutor(max_workers=FRAME_WORKERS,
                                    thread_name_prefix="ev-frame") as pool:
                for f in [pool.submit(work, t) for t in tasks]:
                    f.result()
            with s.lock:
                s.frame_state["state"] = "aborted" if s.abort.is_set() else "done"
        except Exception as exc:
            LOGGER.exception("Frame extraction failed")
            with s.lock:
                s.frame_state["state"] = "error"
                s.frame_state["error"] = str(exc)

    threading.Thread(target=runner, name="ev-frame-runner", daemon=True).start()


# --------------------------------------------------------------------------- #
# On-demand video download
# --------------------------------------------------------------------------- #
def download_video(s: Session, idx: int, kind: str):
    import requests
    key = f"{idx}_{kind}"
    out = s.video_path(idx, kind)
    url = s.url(idx, kind)

    if out.exists() and out.stat().st_size > 0:
        with s.lock:
            s.video_status[key] = {"state": "ready", "pct": 100}
        return
    if not url.lower().startswith("http"):
        with s.lock:
            s.video_status[key] = {"state": "error", "pct": 0, "error": "no url"}
        return

    with s.lock:
        s.video_status[key] = {"state": "downloading", "pct": 0}
    tmp = out.with_suffix(".part")
    try:
        with requests.get(url, stream=True, timeout=(15, 120)) as resp:
            resp.raise_for_status()
            total = int(resp.headers.get("Content-Length") or 0)
            got = 0
            with tmp.open("wb") as f:
                for chunk in resp.iter_content(chunk_size=1 << 18):
                    if not chunk:
                        continue
                    f.write(chunk)
                    got += len(chunk)
                    if total:
                        with s.lock:
                            s.video_status[key] = {"state": "downloading",
                                                   "pct": int(got * 100 / total)}
        tmp.replace(out)
        with s.lock:
            s.video_status[key] = {"state": "ready", "pct": 100}
    except Exception as exc:
        tmp.unlink(missing_ok=True)
        with s.lock:
            s.video_status[key] = {"state": "error", "pct": 0, "error": str(exc)}


# --------------------------------------------------------------------------- #
# Review data
# --------------------------------------------------------------------------- #
def events_payload(s: Session, offset=0, limit=100000):
    out = []
    end = min(offset + limit, len(s.rows))
    with s.lock:
        fstat = dict(s.frame_status)
    for idx in range(offset, end):
        row = s.rows[idx]
        meta = {s.headers[i]: row[i] for i in s.varying_cols if row[i].strip()}
        out.append({
            "idx": idx, "meta": meta, "tags": s.tags.get(idx, []),
            "raw_url": bool(s.url(idx, "raw")),
            "inference_url": bool(s.url(idx, "inference")),
            "raw_frame": fstat.get(f"{idx}_raw", "none"),
            "inference_frame": fstat.get(f"{idx}_inference", "none"),
        })
    return {"total": len(s.rows), "offset": offset, "events": out,
            "columns": s.headers, "constants": s.constant_cols,
            "sortable": [s.headers[i] for i in s.varying_cols],
            "tag_counts": s.tag_counts(),
            "all_columns": [s.headers[i] for i in s.varying_cols],
            "dropped": len(s.headers) - len(s.varying_cols) - len(s.constant_cols),
            "raw_col": s.headers[s.raw_col] if s.raw_col is not None else None,
            "inference_col": s.headers[s.inf_col] if s.inf_col is not None else None}


MAX_TAG_LEN = 48


def clean_tag(name: str) -> str:
    return " ".join(str(name or "").split())[:MAX_TAG_LEN].strip()


def set_tags(s: Session, idx: int, payload: dict):
    if not (0 <= idx < len(s.rows)):
        raise ApiError(404, "no such event")
    with s.lock:
        current = list(s.tags.get(idx, []))
        if "tags" in payload:
            current = []
            for name in payload["tags"] or []:
                name = clean_tag(name)
                if name and name not in current:
                    current.append(name)
        for name in ([payload["add"]] if isinstance(payload.get("add"), str)
                     else payload.get("add") or []):
            name = clean_tag(name)
            if name and name not in current:
                current.append(name)
        for name in ([payload["remove"]] if isinstance(payload.get("remove"), str)
                     else payload.get("remove") or []):
            name = clean_tag(name)
            if name in current:
                current.remove(name)
        if current:
            s.tags[idx] = current
        else:
            s.tags.pop(idx, None)
        s.save_tags()
        return {"idx": idx, "tags": current, "tag_counts": s.tag_counts(),
                "tagged_events": len(s.tags)}


def delete_tag(s: Session, name: str):
    name = clean_tag(name)
    if not name:
        raise ApiError(400, "no tag name given")
    removed = 0
    with s.lock:
        for idx in list(s.tags):
            if name in s.tags[idx]:
                s.tags[idx].remove(name)
                removed += 1
                if not s.tags[idx]:
                    del s.tags[idx]
        s.save_tags()
        return {"name": name, "removed": removed, "tag_counts": s.tag_counts(),
                "tagged_events": len(s.tags)}


def write_tagged_csv(s: Session, filters=None) -> Path:
    """The sheet plus a review_tags column. ``filters`` is a list of buckets to
    include, OR-ed together: "blank" (untagged rows) and "tag:<name>" entries.
    None (or every bucket selected) exports everything."""
    with s.lock:
        tags = dict(s.tags)
    all_opts = {"blank"} | {f"tag:{n}" for names in tags.values() for n in names}
    fset = all_opts if filters is None else (set(filters) & all_opts)

    if fset >= all_opts:
        suffix = ""
        keep = lambda names: True
    else:
        def keep(names):
            if not names:
                return "blank" in fset
            return any(f"tag:{n}" in fset for n in names)
        if len(fset) == 1:
            only = next(iter(fset))
            suffix = "_untagged" if only == "blank" else                 "_" + (re.sub(r"[^A-Za-z0-9._-]+", "_", only[4:]) or "tag")
        else:
            suffix = "_filtered"

    out = s.dir / f"{Path(s.filename).stem}_reviewed{suffix}.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(s.headers + ["review_tags"])
        for i, row in enumerate(s.rows):
            names = tags.get(i, [])
            if keep(names):
                w.writerow(row + ["|".join(names)])
    return out


def state_payload():
    sid = CURRENT["sid"]
    if not sid or sid not in SESSIONS:
        return {"session_id": None}
    s = SESSIONS[sid]
    with s.lock:
        return {"session_id": sid, "filename": s.filename, "rows": len(s.rows),
                "columns": s.headers, "raw_col": s.raw_col,
                "inference_col": s.inf_col,
                "refresh": dict(s.refresh_state), "frames": dict(s.frame_state)}
